from flask import Flask, render_template, request, redirect
import openpyxl

app = Flask(__name__)

# Шлях до Excel-файлу
EXCEL_FILE = "data.xlsx"

# Створення Excel-файлу, якщо його не існує
def create_excel_file():
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "Дані"
        # Додавання заголовків колонок
        sheet.append(["Категорія", "Найменування","Штрихкод","Ціна закупки","Ціна продажу","Кількість"])
        wb.save(EXCEL_FILE)

# Головна сторінка з формою
@app.route("/")
def index():
    return render_template("new_item.html")

# Обробка форми
@app.route("/save", methods=["POST"])
def save_to_excel():
    category = request.form["category"]
    itemname = request.form["itemname"]
    shtrihcode = request.form["shtrihcode"]
    buy_price = request.form["buy_price"]
    sale_price = request.form["sale_price"]
    quntity = request.form["quntity"]

    # Відкриття Excel-файлу і додавання даних
    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb.active
    sheet.append([category, itemname, shtrihcode, buy_price, sale_price, quntity])
    wb.save(EXCEL_FILE)

    return redirect("/")  # Повертає користувача назад на головну сторінку

if __name__ == "__main__":
    create_excel_file()  # Перевіряє, чи існує Excel-файл
    app.run(debug=True)
